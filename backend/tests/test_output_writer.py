"""Tests for output_writer.py and excel_styler.py.

Covers:
  - Output column order matches expected template exactly
  - Output defaults applied
  - Agent_Status styling present in Excel output
  - Processed_Timestamp uses America/Chicago and includes UTC offset
"""

from __future__ import annotations

import io
import re

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.output_writer import build_output_dataframe, write_output_excel
from app.rules_loader import DEFAULT_OUTPUT_COLUMNS
from app.schemas import OutputDefaults, RulesBrain


def _make_brain() -> RulesBrain:
    return RulesBrain(
        output_defaults=OutputDefaults(),
        output_columns=DEFAULT_OUTPUT_COLUMNS[:],
    )


def _sample_rows() -> list[dict]:
    return [
        {
            "Claim_ID": "CLM001",
            "Denial_ID": "DEN001",
            "Reason_Code": "MAT_ATTR_MISMATCH",
            "Primary_Source_Checked": "Material Master",
            "Research_Finding": "Type match confirmed.",
            "Recommended_Next_Action": "No action.",
            "Agent_Status": "Ready for ECC Research Note",
        },
        {
            "Claim_ID": "CLM002",
            "Denial_ID": "DEN002",
            "Reason_Code": "PRICE_VARIANCE",
            "Primary_Source_Checked": "Pricing Data",
            "Research_Finding": "Price outside tolerance.",
            "Recommended_Next_Action": "Review pricing.",
            "Agent_Status": "Needs Manual Review",
        },
        {
            "Claim_ID": "CLM003",
            "Denial_ID": "DEN003",
            "Reason_Code": "MISSING_CONTRACT",
            "Primary_Source_Checked": "Contract Data",
            "Research_Finding": "No matching contract found.",
            "Recommended_Next_Action": "Research contract coverage.",
            "Agent_Status": "Data Missing",
        },
    ]


class TestOutputColumnOrder:
    def test_exact_column_order(self):
        df = build_output_dataframe(_sample_rows(), _make_brain())
        assert list(df.columns) == DEFAULT_OUTPUT_COLUMNS

    def test_no_extra_columns(self):
        df = build_output_dataframe(_sample_rows(), _make_brain())
        assert len(df.columns) == len(DEFAULT_OUTPUT_COLUMNS)


class TestOutputDefaults:
    def test_ecc_update_type_default(self):
        df = build_output_dataframe(_sample_rows(), _make_brain())
        assert all(df["ECC_Update_Type"] == "Research Finding Only")

    def test_financial_posting_allowed_default(self):
        df = build_output_dataframe(_sample_rows(), _make_brain())
        assert all(df["Financial_Posting_Allowed"] == "No")

    def test_pricing_change_allowed_default(self):
        df = build_output_dataframe(_sample_rows(), _make_brain())
        assert all(df["Pricing_Change_Allowed"] == "No")

    def test_existing_values_not_overwritten(self):
        rows = [{"ECC_Update_Type": "Custom", "Agent_Status": "Data Missing"}]
        df = build_output_dataframe(rows, _make_brain())
        assert df.loc[0, "ECC_Update_Type"] == "Custom"


class TestProcessedTimestamp:
    _TIMESTAMP_PATTERN = re.compile(
        r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2} (CDT|CST) \(UTC[+-]\d{2}:\d{2}\)$"
    )

    def test_timestamp_format(self):
        df = build_output_dataframe(_sample_rows(), _make_brain())
        ts = df["Processed_Timestamp"].iloc[0]
        assert self._TIMESTAMP_PATTERN.match(ts), f"Unexpected timestamp format: {ts}"

    def test_timestamp_contains_utc_offset(self):
        df = build_output_dataframe(_sample_rows(), _make_brain())
        ts = df["Processed_Timestamp"].iloc[0]
        assert "UTC" in ts

    def test_timestamp_contains_central_abbrev(self):
        df = build_output_dataframe(_sample_rows(), _make_brain())
        ts = df["Processed_Timestamp"].iloc[0]
        assert "CDT" in ts or "CST" in ts


class TestExcelOutput:
    def _get_workbook(self, rows: list[dict] | None = None) -> object:
        wb_bytes = write_output_excel(rows or _sample_rows(), _make_brain())
        return load_workbook(io.BytesIO(wb_bytes))

    def test_results_sheet_exists(self):
        wb = self._get_workbook()
        assert "Results" in wb.sheetnames

    def test_header_row_matches_template(self):
        wb = self._get_workbook()
        ws = wb["Results"]
        headers = [cell.value for cell in ws[1]]
        assert headers == DEFAULT_OUTPUT_COLUMNS

    def test_agent_status_column_has_fill(self):
        wb = self._get_workbook()
        ws = wb["Results"]
        # Locate Agent_Status column (1-based)
        header_row = [cell.value for cell in ws[1]]
        agent_col_idx = header_row.index("Agent_Status") + 1

        fills_applied = []
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=agent_col_idx)
            fill = cell.fill
            if fill and fill.fill_type == "solid" and fill.fgColor:
                fills_applied.append(fill.fgColor.rgb)

        assert len(fills_applied) > 0, "No Agent_Status fills found in workbook"

    def test_green_fill_for_ready_status(self):
        wb = self._get_workbook()
        ws = wb["Results"]
        header_row = [cell.value for cell in ws[1]]
        agent_col_idx = header_row.index("Agent_Status") + 1

        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=agent_col_idx)
            if status_cell.value == "Ready for ECC Research Note":
                rgb = status_cell.fill.fgColor.rgb
                assert rgb == "FFD6EAD7", f"Expected soft green, got {rgb}"

    def test_yellow_fill_for_needs_review_status(self):
        wb = self._get_workbook()
        ws = wb["Results"]
        header_row = [cell.value for cell in ws[1]]
        agent_col_idx = header_row.index("Agent_Status") + 1

        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=agent_col_idx)
            if status_cell.value == "Needs Manual Review":
                rgb = status_cell.fill.fgColor.rgb
                assert rgb == "FFFFF2CC", f"Expected soft yellow, got {rgb}"

    def test_red_fill_for_data_missing_status(self):
        wb = self._get_workbook()
        ws = wb["Results"]
        header_row = [cell.value for cell in ws[1]]
        agent_col_idx = header_row.index("Agent_Status") + 1

        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=agent_col_idx)
            if status_cell.value == "Data Missing":
                rgb = status_cell.fill.fgColor.rgb
                assert rgb == "FFFCE4E4", f"Expected soft red, got {rgb}"

    def test_check_mark_prepended_to_ready_rows(self):
        wb = self._get_workbook()
        ws = wb["Results"]
        header_row = [cell.value for cell in ws[1]]
        agent_col_idx = header_row.index("Agent_Status") + 1
        finding_col_idx = header_row.index("Research_Finding") + 1

        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=agent_col_idx)
            if status_cell.value == "Ready for ECC Research Note":
                finding_val = ws.cell(row=row_idx, column=finding_col_idx).value or ""
                assert finding_val.startswith("✓"), f"Expected ✓ prefix, got: {finding_val}"

    def test_empty_rows_produces_valid_excel(self):
        wb_bytes = write_output_excel([], _make_brain())
        wb = load_workbook(io.BytesIO(wb_bytes))
        ws = wb["Results"]
        headers = [cell.value for cell in ws[1]]
        assert headers == DEFAULT_OUTPUT_COLUMNS

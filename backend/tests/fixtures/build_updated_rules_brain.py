"""
Build an updated Claims_AI_Rules_Brain Excel file that aligns with the
actual input data files (DenialRecords.xlsx).

Key changes from the original rules brain:
1. Reason codes updated to match input data:
   DIVEST_WRONG_MANUFACTURER  (was DIVEST_VENDOR_MISMATCH)
   DIVEST_PRICE_MISMATCH       (was DIVEST_PRICE_OWNER_MISMATCH)
   DIVEST_CONTRACT_NOT_LOADED  (was DIVEST_CONTRACT_OWNER_MISMATCH)
   DIVEST_CUSTOMER_NOT_ELIGIBLE (was DIVEST_CHARGEBACK_INELIGIBLE)
   DIVEST_TRANSITIONAL_PRICING  (was DIVEST_EFFECTIVE_DATE_GAP)

2. Output_Template updated to 23-column format matching OutputFile.xlsx

3. Status_Color_Rules updated with new Agent_Status values

4. Reason_Code_Aliases added for old→new code mappings
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

FIXTURE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(FIXTURE_DIR, "Claims_AI_Rules_Brain_Updated.xlsx")

# Source rules brain to copy from (user-provided)
SOURCE_PATH = "/Users/adi/Downloads/Claims_AI_Rules_Brain_Renewed.xlsx"


def build_from_scratch():
    """Build a minimal but complete updated rules brain from scratch."""
    wb = Workbook()
    wb.remove(wb.active)

    _add_scenarios(wb)
    _add_join_logic(wb)
    _add_validation_checks(wb)
    _add_field_dictionary(wb)
    _add_reason_code_aliases(wb)
    _add_output_template(wb)
    _add_output_defaults(wb)
    _add_status_color_rules(wb)
    _add_divestiture_business_rules(wb)
    _add_readme(wb)

    wb.save(OUTPUT_PATH)
    print(f"Updated rules brain saved to: {OUTPUT_PATH}")


def _header_row(ws, values, bold=True, fill_color="FFD9D9D9"):
    """Write a styled header row."""
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(bold=bold)
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.fill = fill
        cell.font = font


def _description_row(ws, text):
    """Write a description/title row above the header (per rules brain convention)."""
    ws.append([text])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = Font(bold=True, italic=True)


# ---------------------------------------------------------------------------
# Sheet: Scenarios
# ---------------------------------------------------------------------------
SCENARIOS = [
    # Canonical_Code, Primary_Source, Display_Name, No_Match_Status, Default_Action, Enabled
    ("MAT_ATTR_MISMATCH", "MaterialMasterRecords", "Material Master Records", "Data Missing", "Review material attributes and resubmit.", "Yes"),
    ("CUST_ELIGIBILITY", "CustomerMasterRecords", "Customer Master Records", "Data Missing", "Verify customer eligibility status.", "Yes"),
    ("CONTRACT_MISMATCH", "ContractsData", "Contract Data", "Data Missing", "Review contract assignment.", "Yes"),
    ("MISSING_CONTRACT", "ContractsData", "Contract Data", "Data Missing", "Research contract coverage.", "Yes"),
    ("PRICE_VARIANCE", "PricingData", "Pricing Data", "Data Missing", "Review submitted vs contract price.", "Yes"),
    # Divestiture scenarios — reason codes match input data files
    ("DIVEST_WRONG_MANUFACTURER", "MaterialMasterRecords", "Material Master Records", "Data Missing", "Resubmit to correct manufacturer per divestiture rules.", "Yes"),
    ("DIVEST_PRICE_MISMATCH", "PricingData", "Pricing Data", "Data Missing", "Review pricing under divestiture transitional rules.", "Yes"),
    ("DIVEST_CONTRACT_NOT_LOADED", "ContractsData", "Contract Data", "Data Missing", "Verify contract novation and load status post-divestiture.", "Yes"),
    ("DIVEST_CUSTOMER_NOT_ELIGIBLE", "ContractsData", "Contract Data", "Data Missing", "Validate customer eligibility under current manufacturer contract.", "Yes"),
    ("DIVEST_TRANSITIONAL_PRICING", "PricingData", "Pricing Data", "Data Missing", "Apply transitional pricing agreement per EB_R5.", "Yes"),
]


def _add_scenarios(wb):
    ws = wb.create_sheet("Scenarios")
    _description_row(ws, "Denial Reason Code Scenario Routing — drives primary source lookup and join strategy")
    _header_row(ws, ["Reason_Code", "Primary_Source_File", "Primary_Source_Display", "No_Match_Status", "Default_Recommended_Action", "Enabled"])
    for row in SCENARIOS:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["E"].width = 55


# ---------------------------------------------------------------------------
# Sheet: Join_Logic
# ---------------------------------------------------------------------------
JOIN_LOGIC = [
    # Reason_Code, Driver_Join_Keys (semicolon-sep), Join_Mode, Duplicate_Match_Strategy
    ("MAT_ATTR_MISMATCH", "Material_ID;NDC", "any_key_match", "manual_review"),
    ("CUST_ELIGIBILITY", "Customer_ID", "all_keys_match", "manual_review"),
    ("CONTRACT_MISMATCH", "Contract_ID", "all_keys_match", "manual_review"),
    ("MISSING_CONTRACT", "Customer_ID;Material_ID", "all_keys_match", "allow_multiple_then_validate_active_dates"),
    ("PRICE_VARIANCE", "Contract_ID;Customer_ID;Material_ID", "all_keys_match", "manual_review"),
    ("DIVEST_WRONG_MANUFACTURER", "Material_ID;NDC", "any_key_match", "manual_review"),
    ("DIVEST_PRICE_MISMATCH", "Contract_ID;Material_ID;Customer_ID", "all_keys_match", "manual_review"),
    ("DIVEST_CONTRACT_NOT_LOADED", "Contract_ID;Customer_ID", "any_key_match", "manual_review"),
    ("DIVEST_CUSTOMER_NOT_ELIGIBLE", "Customer_ID;Material_ID", "all_keys_match", "manual_review"),
    ("DIVEST_TRANSITIONAL_PRICING", "Contract_ID;Material_ID", "all_keys_match", "manual_review"),
]


def _add_join_logic(wb):
    ws = wb.create_sheet("Join_Logic")
    _description_row(ws, "Per-scenario join keys and matching strategy")
    _header_row(ws, ["Reason_Code", "Driver_Join_Keys", "Join_Mode", "Duplicate_Match_Strategy"])
    for row in JOIN_LOGIC:
        ws.append(list(row))
    ws.column_dimensions["B"].width = 38


# ---------------------------------------------------------------------------
# Sheet: Validation_Checks
# ---------------------------------------------------------------------------
VALIDATION_CHECKS = [
    # Reason_Code, Field, Operator, Expected_Value_or_Reference, Finding_Note
    # MAT_ATTR_MISMATCH
    ("MAT_ATTR_MISMATCH", "Material_ID", "exists", "", "Material ID not found in Material Master."),
    ("MAT_ATTR_MISMATCH", "Material_Status", "is_not_blank", "", "Material status is missing."),
    # CUST_ELIGIBILITY
    ("CUST_ELIGIBILITY", "Customer_ID", "exists", "", "Customer ID not found in Customer Master."),
    ("CUST_ELIGIBILITY", "Eligibility_Status", "equals", "Eligible", "Customer is not eligible for chargeback."),
    # CONTRACT_MISMATCH
    ("CONTRACT_MISMATCH", "Contract_ID", "exists", "", "Contract ID not found in Contract Data."),
    ("CONTRACT_MISMATCH", "Contract_Status", "equals", "Active", "Contract is not in Active status."),
    # PRICE_VARIANCE
    ("PRICE_VARIANCE", "Expected_Price", "is_not_blank", "", "Expected price not found in Pricing Data."),
    ("PRICE_VARIANCE", "Submitted_Price", "is_not_blank", "", "Submitted price is blank on denial record."),
    ("PRICE_VARIANCE", "Submitted_Price", "price_difference", "Expected_Price", "Price variance exceeds tolerance."),
    # DIVEST_WRONG_MANUFACTURER
    ("DIVEST_WRONG_MANUFACTURER", "Material_ID", "exists", "", "Material not found in Material Master."),
    ("DIVEST_WRONG_MANUFACTURER", "Divestiture_Flag", "equals", "Yes", "Material is not flagged as divested."),
    ("DIVEST_WRONG_MANUFACTURER", "Divestiture_Effective_Date", "is_not_blank", "", "Divestiture effective date is missing."),
    ("DIVEST_WRONG_MANUFACTURER", "Invoice_Date", "is_not_blank", "", "Invoice date is missing on denial record."),
    ("DIVEST_WRONG_MANUFACTURER", "Submitted_Manufacturer", "is_not_blank", "", "Submitted manufacturer is blank."),
    # DIVEST_PRICE_MISMATCH
    ("DIVEST_PRICE_MISMATCH", "Expected_Price", "is_not_blank", "", "Expected price not found in Pricing Data."),
    ("DIVEST_PRICE_MISMATCH", "Transitional_Pricing_Flag", "is_not_blank", "", "Transitional pricing flag missing from pricing record."),
    ("DIVEST_PRICE_MISMATCH", "Submitted_Price", "price_difference", "Expected_Price", "Price mismatch under divestiture pricing rules."),
    # DIVEST_CONTRACT_NOT_LOADED
    ("DIVEST_CONTRACT_NOT_LOADED", "Contract_ID", "exists", "", "Contract not found in Contract Data post-divestiture."),
    ("DIVEST_CONTRACT_NOT_LOADED", "Contract_Status", "is_not_blank", "", "Contract status is missing."),
    ("DIVEST_CONTRACT_NOT_LOADED", "Novation_Flag", "is_not_blank", "", "Novation flag missing — cannot determine contract validity."),
    # DIVEST_CUSTOMER_NOT_ELIGIBLE
    ("DIVEST_CUSTOMER_NOT_ELIGIBLE", "Customer_ID", "exists", "", "Customer not found in Contract Data."),
    ("DIVEST_CUSTOMER_NOT_ELIGIBLE", "Chargeback_Eligible_Flag", "equals", "Yes", "Customer is not chargeback-eligible under current manufacturer contract."),
    ("DIVEST_CUSTOMER_NOT_ELIGIBLE", "Contract_Assignment_Status", "is_not_blank", "", "Contract assignment status is missing."),
    # DIVEST_TRANSITIONAL_PRICING
    ("DIVEST_TRANSITIONAL_PRICING", "Transitional_Pricing_Flag", "equals", "Yes", "Transitional pricing flag not set — may not qualify for transitional pricing."),
    ("DIVEST_TRANSITIONAL_PRICING", "Transition_End_Date", "is_not_blank", "", "Transition end date is missing."),
    ("DIVEST_TRANSITIONAL_PRICING", "Invoice_Date", "between_dates", "Divestiture_Effective_Date,Transition_End_Date", "Invoice date falls outside the transitional pricing window."),
]


def _add_validation_checks(wb):
    ws = wb.create_sheet("Validation_Checks")
    _description_row(ws, "Per-scenario field-level validation rules — drives Research_Finding and Agent_Status")
    _header_row(ws, ["Reason_Code", "Field_or_Field_Group", "Operator", "Expected_Value_or_Reference", "Finding_Note"])
    for row in VALIDATION_CHECKS:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["E"].width = 60


# ---------------------------------------------------------------------------
# Sheet: Field_Dictionary
# ---------------------------------------------------------------------------
FIELD_DICT = [
    # Canonical_Field, Applies_To_Source, Accepted_Aliases, Data_Type
    ("Claim_ID", "DenialRecords", "Claim_ID;ClaimID;Claim ID", "string"),
    ("Denial_ID", "DenialRecords", "Denial_ID;DenialID;Denial ID", "string"),
    ("Reason_Code", "DenialRecords", "Reason_Code;ReasonCode;Reason Code", "string"),
    ("Customer_ID", "DenialRecords;CustomerMasterRecords;ContractsData;PricingData", "Customer_ID;CustomerID;Customer ID", "string"),
    ("Material_ID", "DenialRecords;MaterialMasterRecords;ContractsData;PricingData", "Material_ID;MaterialID;Material ID", "string"),
    ("NDC", "DenialRecords;MaterialMasterRecords;ContractsData;PricingData", "NDC;Ndc;ndc", "string"),
    ("Contract_ID", "DenialRecords;ContractsData;PricingData", "Contract_ID;ContractID;Contract ID", "string"),
    ("Invoice_Date", "DenialRecords", "Invoice_Date;InvoiceDate;Invoice Date", "date"),
    ("Submission_Date", "DenialRecords", "Submission_Date;SubmissionDate;Submission Date", "date"),
    ("Submitted_Manufacturer", "DenialRecords", "Submitted_Manufacturer;SubmittedManufacturer", "string"),
    ("Expected_Manufacturer", "DenialRecords", "Expected_Manufacturer;ExpectedManufacturer", "string"),
    ("Submitted_Price", "DenialRecords;PricingData", "Submitted_Price;SubmittedPrice;Submitted Price", "decimal"),
    ("Expected_Price", "PricingData", "Expected_Price;ExpectedPrice;Expected Price", "decimal"),
    ("Divestiture_Related_Flag", "DenialRecords", "Divestiture_Related_Flag;DivestitureRelatedFlag", "string"),
    ("Inventory_Purchased_From", "DenialRecords", "Inventory_Purchased_From;InventoryPurchasedFrom", "string"),
    ("Inventory_Purchase_Date", "DenialRecords", "Inventory_Purchase_Date;InventoryPurchaseDate", "date"),
    ("Sale_Date", "DenialRecords", "Sale_Date;SaleDate", "date"),
    # Material Master fields
    ("Divestiture_Flag", "MaterialMasterRecords", "Divestiture_Flag;DivestitureFlag", "string"),
    ("Divestiture_Effective_Date", "MaterialMasterRecords", "Divestiture_Effective_Date;DivestitureEffectiveDate", "date"),
    ("Prior_Manufacturer", "MaterialMasterRecords", "Prior_Manufacturer;PriorManufacturer", "string"),
    ("Current_Manufacturer", "MaterialMasterRecords", "Current_Manufacturer;CurrentManufacturer", "string"),
    ("Transitional_Pricing_Flag", "MaterialMasterRecords;PricingData", "Transitional_Pricing_Flag;TransitionalPricingFlag", "string"),
    ("Transition_End_Date", "MaterialMasterRecords;PricingData", "Transition_End_Date;TransitionEndDate", "date"),
    ("Trade_Letter_Override_Flag", "MaterialMasterRecords;ContractsData", "Trade_Letter_Override_Flag;TradeLetter Override Flag", "string"),
    ("Material_Status", "MaterialMasterRecords", "Material_Status;MaterialStatus", "string"),
    # Contract fields
    ("Contract_Status", "ContractsData", "Contract_Status;ContractStatus", "string"),
    ("Contract_Start_Date", "ContractsData", "Contract_Start_Date;ContractStartDate", "date"),
    ("Contract_End_Date", "ContractsData", "Contract_End_Date;ContractEndDate", "date"),
    ("Contract_Price", "ContractsData", "Contract_Price;ContractPrice", "decimal"),
    ("Novation_Flag", "ContractsData", "Novation_Flag;NovationFlag", "string"),
    ("Successor_Contract_ID", "ContractsData", "Successor_Contract_ID;SuccessorContractID", "string"),
    ("Prior_Contract_ID", "ContractsData", "Prior_Contract_ID;PriorContractID", "string"),
    ("Contract_Owner", "ContractsData", "Contract_Owner;ContractOwner", "string"),
    ("Contract_Owner_Vendor_ID", "ContractsData", "Contract_Owner_Vendor_ID;ContractOwnerVendorID", "string"),
    ("Contract_Assignment_Status", "ContractsData", "Contract_Assignment_Status;ContractAssignmentStatus", "string"),
    ("Chargeback_Eligible_Flag", "ContractsData;CustomerMasterRecords", "Chargeback_Eligible_Flag;ChargebackEligibleFlag", "string"),
    # Customer Master fields
    ("Eligibility_Status", "CustomerMasterRecords", "Eligibility_Status;EligibilityStatus;Eligible", "string"),
    ("Customer_Status", "CustomerMasterRecords", "Customer_Status;CustomerStatus", "string"),
    ("GPO_Alignment", "CustomerMasterRecords", "GPO_Alignment;GPOAlignment", "string"),
    # Pricing fields
    ("Price_Variance", "PricingData", "Price_Variance;PriceVariance", "decimal"),
    ("Pricing_Effective_Date", "PricingData", "Pricing_Effective_Date;PricingEffectiveDate", "date"),
    ("Pricing_Expiration_Date", "PricingData", "Pricing_Expiration_Date;PricingExpirationDate", "date"),
]


def _add_field_dictionary(wb):
    ws = wb.create_sheet("Field_Dictionary")
    _description_row(ws, "Canonical field name dictionary — maps raw column names to canonical names for all sources")
    _header_row(ws, ["Canonical_Field", "Applies_To_Source", "Accepted_Aliases", "Data_Type"])
    for row in FIELD_DICT:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55


# ---------------------------------------------------------------------------
# Sheet: Reason_Code_Aliases
# ---------------------------------------------------------------------------
REASON_CODE_ALIASES = [
    # Canonical, Accepted_Aliases (semicolon-separated)
    ("MAT_ATTR_MISMATCH", "MAT_ATTR_MISMATCH;mat_attr_mismatch;Material Attribute Mismatch;Material-Attr-Mismatch"),
    ("CUST_ELIGIBILITY", "CUST_ELIGIBILITY;cust_eligibility;Customer Eligibility;Customer-Eligibility"),
    ("CONTRACT_MISMATCH", "CONTRACT_MISMATCH;contract_mismatch;Contract Mismatch;Contract-Mismatch"),
    ("MISSING_CONTRACT", "MISSING_CONTRACT;missing_contract;Missing Contract;Missing-Contract"),
    ("PRICE_VARIANCE", "PRICE_VARIANCE;price_variance;Price Variance;Price-Variance"),
    # Divestiture codes — canonical = input-file codes; aliases include old rules brain names
    ("DIVEST_WRONG_MANUFACTURER", "DIVEST_WRONG_MANUFACTURER;divest_wrong_manufacturer;DIVEST_VENDOR_MISMATCH;Divest Wrong Manufacturer;Divestiture Wrong Manufacturer"),
    ("DIVEST_PRICE_MISMATCH", "DIVEST_PRICE_MISMATCH;divest_price_mismatch;DIVEST_PRICE_OWNER_MISMATCH;Divest Price Mismatch;Divestiture Price Mismatch"),
    ("DIVEST_CONTRACT_NOT_LOADED", "DIVEST_CONTRACT_NOT_LOADED;divest_contract_not_loaded;DIVEST_CONTRACT_OWNER_MISMATCH;Divest Contract Not Loaded"),
    ("DIVEST_CUSTOMER_NOT_ELIGIBLE", "DIVEST_CUSTOMER_NOT_ELIGIBLE;divest_customer_not_eligible;DIVEST_CHARGEBACK_INELIGIBLE;Divest Customer Not Eligible"),
    ("DIVEST_TRANSITIONAL_PRICING", "DIVEST_TRANSITIONAL_PRICING;divest_transitional_pricing;DIVEST_EFFECTIVE_DATE_GAP;Divest Transitional Pricing"),
]


def _add_reason_code_aliases(wb):
    ws = wb.create_sheet("Reason_Code_Aliases")
    _description_row(ws, "Canonical reason code normalisation map — all aliases resolve to one canonical code")
    _header_row(ws, ["Canonical_Reason_Code", "Accepted_Aliases"])
    for row in REASON_CODE_ALIASES:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80


# ---------------------------------------------------------------------------
# Sheet: Output_Template — 23 columns matching OutputFile.xlsx
# ---------------------------------------------------------------------------
OUTPUT_COLUMNS = [
    "Claim_ID",
    "Denial_ID",
    "Reason_Code",
    "Divestiture_Related_Flag",
    "Invoice_Date",
    "Submission_Date",
    "Submitted_Manufacturer",
    "Expected_Manufacturer",
    "Ownership_Determination",
    "Transition_Period_Flag",
    "Primary_Source_Checked",
    "Secondary_Source_Checked",
    "Data_Validation_Result",
    "Research_Finding",
    "Discrepancy_Details",
    "Denial_Decision",
    "Resubmission_Recommended",
    "Recommended_Next_Action",
    "ECC_Update_Type",
    "Financial_Posting_Allowed",
    "Pricing_Change_Allowed",
    "Agent_Status",
    "Processed_Timestamp",
]


def _add_output_template(wb):
    ws = wb.create_sheet("Output_Template")
    _description_row(ws, "Output Excel column order — 23 columns (11 core + 12 divestiture). ECC_Update_Type always Research Finding Only.")
    _header_row(ws, ["Column_Name", "Notes"])
    notes = {
        "ECC_Update_Type": "Always: Research Finding Only",
        "Financial_Posting_Allowed": "Always: No",
        "Pricing_Change_Allowed": "Always: No",
        "Ownership_Determination": "Derived: Invoice_Date vs Divestiture_Effective_Date",
        "Transition_Period_Flag": "Derived: Yes if within transitional pricing window",
        "Denial_Decision": "Derived: Resubmission Candidate or Acceptable Denial",
        "Resubmission_Recommended": "Derived: Yes or No",
        "Data_Validation_Result": "Derived: Mismatch or Missing when validation fails",
        "Secondary_Source_Checked": "Trade Letter when EB_R7 override applies",
    }
    for col in OUTPUT_COLUMNS:
        ws.append([col, notes.get(col, "")])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55


# ---------------------------------------------------------------------------
# Sheet: Output_Defaults
# ---------------------------------------------------------------------------
def _add_output_defaults(wb):
    ws = wb.create_sheet("Output_Defaults")
    _description_row(ws, "Default values applied to every output row")
    _header_row(ws, ["Key", "Value"])
    ws.append(["ECC_Update_Type", "Research Finding Only"])
    ws.append(["Financial_Posting_Allowed", "No"])
    ws.append(["Pricing_Change_Allowed", "No"])


# ---------------------------------------------------------------------------
# Sheet: Status_Color_Rules
# ---------------------------------------------------------------------------
def _add_status_color_rules(wb):
    ws = wb.create_sheet("Status_Color_Rules")
    _description_row(ws, "Agent_Status values and their Excel fill colors")
    _header_row(ws, ["Agent_Status", "Fill_Color_HEX", "Meaning"])
    ws.append(["Ready for Resubmission Review", "#D6EAD7", "All checks passed — denial likely correctable via resubmission"])
    ws.append(["Closed - Research Complete", "#D0E4F7", "Denial confirmed as valid — no resubmission recommended"])
    ws.append(["Needs Manual Review", "#FFF2CC", "One or more validation checks failed"])
    ws.append(["Data Missing", "#FCE4E4", "Required source record not found"])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["C"].width = 55


# ---------------------------------------------------------------------------
# Sheet: Divestiture_Business_Rules (EB_R1–EB_R7)
# ---------------------------------------------------------------------------
DIVEST_RULES = [
    ("EB_R1", "Ownership Determination",
     "Any divestiture-related denial",
     "IF Invoice_Date < Divestiture_Effective_Date THEN owner = Prior_Manufacturer (Pfizer); ELSE owner = Current_Manufacturer (Viatris)",
     "Invoice_Date;Divestiture_Effective_Date;Prior_Manufacturer;Current_Manufacturer",
     "Determine ownership using INVOICE DATE, not submission date. Document only."),
    ("EB_R2", "Inventory Position",
     "Inventory purchased pre-divest, sold post-divest",
     "IF purchased from Prior_Manufacturer pre-divest AND sold post-divest THEN apply Transitional Pricing Rule",
     "Material_ID;Invoice_Date;Transitional_Pricing_Flag;Inventory_Purchased_From",
     "Document inventory position. Do not adjust pricing or eligibility."),
    ("EB_R3", "Contract Validity",
     "Contract referenced on divestiture denial",
     "IF Novation_Flag = Yes THEN use Successor_Contract_ID; ELSE Prior Contract may still apply per Trade Letter",
     "Contract_Owner;Novation_Flag;Successor_Contract_ID",
     "Check Novation_Flag and Successor_Contract_ID. Document only."),
    ("EB_R4", "Chargeback Submission",
     "Any divestiture chargeback denial",
     "Claims must be submitted to Manufacturer who OWNS product at TIME OF SALE (Invoice_Date). IF Submitted_Manufacturer != expected owner THEN WRONG MANUFACTURER.",
     "Invoice_Date;Submitted_Manufacturer;Expected_Manufacturer",
     "Compare submitted to expected manufacturer at time of sale. Document only."),
    ("EB_R5", "Pricing Continuity",
     "Pricing on divestiture denial",
     "IF Transitional_Pricing_Flag = Yes AND Invoice_Date <= Transition_End_Date THEN honor prior pricing; ELSE apply current manufacturer pricing.",
     "Transitional_Pricing_Flag;Transition_End_Date;Invoice_Date",
     "Check Transitional_Pricing_Flag and Transition_End_Date. Document only. Do not change pricing."),
    ("EB_R6", "Customer Eligibility",
     "Customer on divestiture denial",
     "Customer must be eligible under CURRENT Manufacturer contract. IF Chargeback_Eligible_Flag != Yes THEN CUSTOMER NOT ELIGIBLE.",
     "Customer_ID;Contract_Owner_Vendor_ID;Eligibility_Status;Chargeback_Eligible_Flag",
     "Validate customer alignment to current manufacturer. Document only."),
    ("EB_R7", "Trade Letter Override",
     "Trade_Letter_Override_Flag = Yes (in Material Master or Contracts Data)",
     "Trade Letter overrides ALL rules. Follow Trade Letter instructions exactly.",
     "Trade_Letter_Override_Flag",
     "Never override Trade Letter logic. Document the Trade Letter reference in the finding."),
]


def _add_divestiture_business_rules(wb):
    ws = wb.create_sheet("Divestiture_Business_Rules")
    _description_row(ws, "Established Brands (Pfizer→Viatris) divestiture overlay rules — EB_R1 through EB_R7")
    _header_row(ws, ["Rule_ID", "Rule_Name", "Trigger", "Key_Logic", "Fields_Checked", "Agent_Action"])
    for row in DIVEST_RULES:
        ws.append(list(row))
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 55


# ---------------------------------------------------------------------------
# Sheet: README
# ---------------------------------------------------------------------------
def _add_readme(wb):
    ws = wb.create_sheet("README")
    ws.append(["Claims AI Rules Brain — Updated for Divestiture Additions"])
    ws.append([""])
    ws.append(["This workbook is the single uploadable logic brain for claims denial validation."])
    ws.append(["Backend reads this file to route scenarios, join data, run checks, and generate output."])
    ws.append([""])
    ws.append(["Key update: Divestiture reason codes now match input data files (DenialRecords.xlsx):"])
    ws.append(["  DIVEST_WRONG_MANUFACTURER  (was DIVEST_VENDOR_MISMATCH)"])
    ws.append(["  DIVEST_PRICE_MISMATCH       (was DIVEST_PRICE_OWNER_MISMATCH)"])
    ws.append(["  DIVEST_CONTRACT_NOT_LOADED  (was DIVEST_CONTRACT_OWNER_MISMATCH)"])
    ws.append(["  DIVEST_CUSTOMER_NOT_ELIGIBLE (was DIVEST_CHARGEBACK_INELIGIBLE)"])
    ws.append(["  DIVEST_TRANSITIONAL_PRICING  (was DIVEST_EFFECTIVE_DATE_GAP)"])
    ws.append([""])
    ws.append(["Output expanded to 23 columns matching OutputFile.xlsx."])
    ws.append(["Agent_Status values updated: Ready for Resubmission Review / Closed - Research Complete / Needs Manual Review / Data Missing"])


if __name__ == "__main__":
    build_from_scratch()

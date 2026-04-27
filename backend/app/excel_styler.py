"""Excel styler.

Applies openpyxl formatting to the output workbook:
  - Header row: bold, light grey background.
  - Agent_Status column: conditional fill color based on value.
  - Auto-width for all columns (approximate).

Colors are professional and muted — not loud.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Color palette (ARGB hex, openpyxl format: AARRGGBB)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9D9")   # light grey
HEADER_FONT = Font(bold=True)

# Agent_Status fills – subtle, professional
STATUS_FILLS = {
    "Ready for ECC Research Note": PatternFill(fill_type="solid", fgColor="FFD6EAD7"),  # soft green
    "Needs Manual Review":         PatternFill(fill_type="solid", fgColor="FFFFF2CC"),  # soft yellow
    "Data Missing":                PatternFill(fill_type="solid", fgColor="FFFCE4E4"),  # soft red
}

CHECK_MARK = "✓"   # ✓  – prepended to finding for clean rows


def _col_index(worksheet: Worksheet, header: str) -> int | None:
    """Return 1-based column index for the given header label, or None."""
    for idx, cell in enumerate(worksheet[1], start=1):
        if str(cell.value or "").strip() == header:
            return idx
    return None


def style_workbook(
    workbook: Workbook,
    worksheet: Worksheet,
    df: pd.DataFrame,
) -> None:
    """Apply all formatting in-place to *worksheet*."""
    _style_header(worksheet)
    _style_agent_status(worksheet, df)
    _auto_width(worksheet, df)


def _style_header(worksheet: Worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _style_agent_status(worksheet: Worksheet, df: pd.DataFrame) -> None:
    if "Agent_Status" not in df.columns:
        return
    col_idx = _col_index(worksheet, "Agent_Status")
    if col_idx is None:
        return

    finding_col_idx = _col_index(worksheet, "Research_Finding")

    # Row 1 is the header; data starts at row 2
    for row_num, status_val in enumerate(df["Agent_Status"], start=2):
        status_str = str(status_val or "").strip()
        fill = STATUS_FILLS.get(status_str)
        cell = worksheet.cell(row=row_num, column=col_idx)
        if fill:
            cell.fill = fill

        # Optionally prepend ✓ to Research_Finding for successful rows
        if status_str == "Ready for ECC Research Note" and finding_col_idx:
            finding_cell = worksheet.cell(row=row_num, column=finding_col_idx)
            raw = str(finding_cell.value or "").strip()
            if raw and not raw.startswith(CHECK_MARK):
                finding_cell.value = f"{CHECK_MARK} {raw}"


def _auto_width(worksheet: Worksheet, df: pd.DataFrame) -> None:
    """Set approximate column widths based on header and data content."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if not df.empty else 0,
        )
        # Cap width; add small padding
        worksheet.column_dimensions[col_letter].width = min(max_len + 4, 60)
